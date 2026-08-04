# ---------------------------------------------------------------------------
# build_export_xlsx.py — בונה את קובץ הייצוא לאקסל של ההדגמה.
#
# מה זה
# -----
# בגרסת השרת המלאה כפתור "ייצוא אקסל" הפיק קובץ מהנתונים החיים. בגרסת
# הפורטפוליו הסטטית (GitHub Pages) אין שרת, ולכן הקובץ מיוצר מראש מתוך
# data/*.json (אותם קבצים שה-JS טוען) ונשמר כ-data/export.xlsx. כפתורי
# הייצוא באתר מורידים בדיוק את הקובץ הזה.
#
# פרטיות
# ------
# הקובץ נבנה אך ורק מ-data/*.json — נתונים סינתטיים שהוגרלו ב-seed קבוע.
# אין בו נתוני איוש, דמוגרפיה או מועמדים, ולא נגזר ממנו שום נתון תפעולי.
# זה הקובץ היחיד מסוג *.xlsx שמותר במאגר הציבורי (ראו החריג ב-.gitignore).
#
# מבנה הטבלה
# ----------
# גיליון 1: שורה לתחנה — זיהוי, סה"כ משרות פתוחות, סטטוס, ועמודה לכל אחת
# משש משפחות המקצוע. גיליון 2: סיכום (כמויות, התפלגות פעמון, לפי סטטוס,
# לפי מחוז, לפי משפחת מקצוע).
# ---------------------------------------------------------------------------
import json
import statistics
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_THIN = Side(style="thin", color="C8CDD6")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_GRP_FILL = PatternFill("solid", fgColor="111827")
_SUB_FILL = PatternFill("solid", fgColor="374151")
_WHITE_B = Font(name="David", bold=True, color="FFFFFF", size=11)
_CELL = Font(name="David", size=11)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# הספים חייבים להתאים ל-generate_demo_data.THRESHOLDS. הם נקראים מ-legend.json
# בזמן הבנייה כדי שלא יהיו שני מקורות אמת שיכולים להיפרד.
_STATUS_ORDER = ["קריטי", "דחוף", "בינוני", "תקין"]
_STATUS_COLOR = {"קריטי": "DC2626", "דחוף": "EA580C", "בינוני": "EAB308", "תקין": "16A34A"}


def _load(data_dir):
    data_dir = Path(data_dir)
    stations = json.loads((data_dir / "stations.json").read_text(encoding="utf-8"))
    # סדר משפחות המקצוע נלקח מהתחנה הראשונה — הוא זהה בכל התחנות, ולכן אין
    # צורך (ואסור) לשכפל אותו כרשימה קשיחה כאן.
    families = [(r["key"], r["label"]) for r in stations[0]["roles"]]
    return stations, families


def _sheet_stations(wb, stations, families):
    ws = wb.active
    ws.title = "משרות פתוחות"
    ws.sheet_view.rightToLeft = True

    groups = [
        ("זיהוי", ["מחוז", "מרחב", "תחנה"]),
        ('סה"כ תחנה', ["משרות פתוחות", "סטטוס", "מקצוע מוביל"]),
        ("פילוח לפי מקצוע", [label for _, label in families]),
    ]

    col = 1
    for gname, subs in groups:
        start = col
        for sub in subs:
            ws.cell(2, col, sub)
            col += 1
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=col - 1)
        ws.cell(1, start, gname)
    ncols = col - 1
    for c in range(1, ncols + 1):
        g = ws.cell(1, c); g.fill = _GRP_FILL; g.font = _WHITE_B; g.alignment = _CENTER; g.border = _BORDER
        s = ws.cell(2, c); s.fill = _SUB_FILL; s.font = _WHITE_B; s.alignment = _CENTER; s.border = _BORDER

    rows = sorted(stations, key=lambda s: (s["district"], s["area"], s["name"]))
    r = 3
    for s in rows:
        by_key = {x["key"]: x["open"] for x in s["roles"]}
        top = s.get("top_role")
        vals = [s["district"], s["area"], s["name"],
                s["open_positions"], s["status"], top["label"] if top else "—"]
        vals += [by_key[key] for key, _ in families]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v); cell.font = _CELL; cell.alignment = _CENTER; cell.border = _BORDER
        # עמודות 4-5 (סה"כ + סטטוס) נצבעות בצבע הסטטוס — כך שאפשר לסרוק את
        # הגיליון בעין ולראות את אותה מפת חום שרואים באתר.
        for c in (4, 5):
            cell = ws.cell(r, c)
            cell.fill = PatternFill("solid", fgColor=_STATUS_COLOR.get(s["status"], "9CA3AF"))
            cell.font = Font(name="David", color="FFFFFF", size=11, bold=(c == 4))
        r += 1

    widths = [11, 11, 20, 14, 10, 14] + [12] * len(families)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "D3"
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 30


def _sheet_summary(wb, stations, families):
    ws = wb.create_sheet("סיכום")
    ws.sheet_view.rightToLeft = True

    def H(txt, row):
        ws.cell(row, 1, txt).font = Font(name="David", bold=True, size=13)
        return row + 1

    def HR(row, cols):
        for j, t in enumerate(cols, 1):
            c = ws.cell(row, j, t); c.fill = _GRP_FILL; c.font = _WHITE_B; c.alignment = _CENTER; c.border = _BORDER
        return row + 1

    def DR(row, vals, fill=None):
        for j, v in enumerate(vals, 1):
            c = ws.cell(row, j, v); c.font = _CELL; c.alignment = _CENTER; c.border = _BORDER
            if fill and j == 1:
                c.fill = PatternFill("solid", fgColor=fill)
                c.font = Font(name="David", bold=True, color="FFFFFF", size=11)
        return row + 1

    opens = sorted(s["open_positions"] for s in stations)
    r = 1
    r = H("כמויות", r)
    r = HR(r, ["מדד", "כמות"])
    r = DR(r, ["מחוזות", len({s["district"] for s in stations})])
    r = DR(r, ["מרחבים", len({(s["district"], s["area"]) for s in stations})])
    r = DR(r, ["תחנות", len(stations)])
    r = DR(r, ['סה"כ משרות פתוחות', sum(opens)])
    r += 1

    r = H("התפלגות משרות פתוחות לתחנה (עקומת פעמון)", r)
    ws.cell(r, 1, f"ממוצע {statistics.mean(opens):.1f}  ·  חציון {statistics.median(opens):.0f}  ·  "
                  f"סטיית תקן {statistics.pstdev(opens):.1f}").font = _CELL
    r += 1
    r = HR(r, ["טווח", "תחנות", "גרף"])
    buckets = Counter(int(v // 4) * 4 for v in opens)
    for k in range(0, max(opens) + 1, 4):
        n = buckets.get(k, 0)
        r = DR(r, [f"{k}-{k + 3}", n, "█" * n])
    r += 1

    r = H("פילוח לפי סטטוס", r)
    r = HR(r, ["סטטוס", "תחנות", 'סה"כ משרות פתוחות'])
    sc = Counter(s["status"] for s in stations)
    so = defaultdict(int)
    for s in stations:
        so[s["status"]] += s["open_positions"]
    for stt in _STATUS_ORDER:
        r = DR(r, [stt, sc.get(stt, 0), so.get(stt, 0)], fill=_STATUS_COLOR[stt])
    r += 1

    r = H("פילוח לפי מחוז", r)
    r = HR(r, ["מחוז", "תחנות", "מרחבים", "משרות פתוחות", "ממוצע לתחנה"])
    byd = defaultdict(lambda: {"n": 0, "areas": set(), "open": 0})
    for s in stations:
        d = byd[s["district"]]
        d["n"] += 1; d["areas"].add(s["area"]); d["open"] += s["open_positions"]
    for dist, d in sorted(byd.items(), key=lambda x: -x[1]["open"]):
        r = DR(r, [dist, d["n"], len(d["areas"]), d["open"], round(d["open"] / d["n"], 1)])
    r += 1

    r = H("פילוח לפי משפחת מקצוע", r)
    r = HR(r, ["משפחה", "משרות פתוחות", "אחוז מהסך", "גרף"])
    fam = defaultdict(int)
    for s in stations:
        for x in s["roles"]:
            fam[x["key"]] += x["open"]
    total = sum(fam.values()) or 1
    for key, label in sorted(families, key=lambda f: -fam[f[0]]):
        r2 = DR(r, [label, fam[key], fam[key] / total, "█" * round(fam[key] / total * 50)])
        ws.cell(r, 3).number_format = "0.0%"; r = r2

    for i, w in enumerate([18, 16, 14, 16, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build(data_dir, out_path=None):
    """בונה את data/export.xlsx מתוך קובצי ה-JSON הסינתטיים. מחזיר את הנתיב."""
    data_dir = Path(data_dir)
    out_path = Path(out_path) if out_path else data_dir / "export.xlsx"
    stations, families = _load(data_dir)
    wb = Workbook()
    _sheet_stations(wb, stations, families)
    _sheet_summary(wb, stations, families)
    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    p = build(Path(__file__).resolve().parent.parent / "data")
    print(f"נכתב {p}")
